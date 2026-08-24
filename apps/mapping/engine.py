"""
Template Mapping Engine
========================
This is the piece that lets the dashboard never depend on fixed column
positions. Every incoming project format (BPW, Latvia, SSH, HOR, ...) gets
one ProjectTemplate whose `config` JSON says, per STANDARD_SCHEMA field,
*how* to pull the value out of that project's particular workbook shape.

Real-world project trackers are not always one flat table with a header row
(the two files this was built against - BPW and Latvia - keep their live
KPIs in a "Project Summary" sheet with fixed label/value cell pairs, and a
row-per-file "Calculation" sheet for image/record totals). So each field's
rule declares a `mode`:

  "cell"        -> a single fixed cell, e.g. {"sheet": "Project Summary", "cell": "B2"}
  "sum_column"  -> sum every value under a column header on a flat sheet,
                   e.g. {"sheet": "Calculation", "column": "Total Image"}
  "header"      -> classic flat-table lookup: find this column header
                   anywhere in the sheet's header row and read a specific
                   row/first value, e.g. {"sheet": "Delivery Plan", "column": "Delivered Records"}
  "static"      -> a fixed value that doesn't come from the file at all
                   (used for `branch`, since the workbook doesn't state it),
                   e.g. {"value": "TDM"}
  "row_count"   -> counts real data rows on a flat sheet, e.g.
                   {"sheet": "Inventory", "id_column": "S.No"} - useful for
                   `total_batches`/`promoted` when the true count is "every
                   row", not a number already sitting in a cell. Optional
                   `where_column` counts only rows where that column is
                   filled in (e.g. "promoted" = rows with a Shipment date);
                   `where_column_blank` is the inverse (e.g.
                   "batches_being_keyed" = rows with no Shipment date yet).
  "sum_multi"   -> sums this SAME field across several sheets/rules, e.g. a
                   project whose batches are split across two Inventory
                   sheets with no single sheet holding the true total:
                   {"mode": "sum_multi", "sources": [
                     {"mode": "row_count", "sheet": "Inventory_01"},
                     {"mode": "row_count", "sheet": "Inventory_02"}]}
                   `sources` takes any rule above (not "sum_multi" itself).

`inventory_rows` (the per-row Inventory Tracker detail feeding
InventoryItem/the Milestone "PctBy X%" columns) can likewise be either one
rule (single sheet - the common case) or a LIST of rules, one per sheet,
whose rows are unioned - e.g. BV_Inventory_01 + BV_Inventory_02, each with
its own column names for the same logical fields.

Every pandas-backed rule ("sum_column", "header", "row_count", and
inventory_rows") accepts an optional `header_row` (1-indexed, default 1) for
sheets whose real column headers aren't on row 1 - e.g. a row 1 that holds
target/summary numbers above the actual header row, e.g.
{"mode": "sum_column", "sheet": "Inventory", "column": "# of Images", "header_row": 2}

Adding project #5 - even one with a differently-split Inventory - never
touches this file or the dashboard code, only a new JSON template.

Performance note: a large workbook (tens of MB, tens of thousands of rows)
used to get parsed from scratch for every single field - once to validate
headers, once more for cell-mode lookups, and once per sum_column/header/
inventory_rows rule via pandas. On a real 29MB file that added up to several
minutes. This version opens the workbook exactly once in read-only mode for
all cell-mode work, and caches each sheet's DataFrame the first time any
pandas-based rule touches it, so a template with several rules pointing at
the same sheet only pays the parse cost once.
"""
from __future__ import annotations

import datetime
from dataclasses import dataclass, field

import openpyxl
import pandas as pd

STANDARD_SCHEMA = [
    "project_name", "branch", "start_date", "end_date",
    "target", "delivered", "images",
    "language", "vendor", "event_type", "ocr_status",
    "total_batches", "batches_being_keyed", "promoted",
]

REQUIRED_FIELDS = ["project_name", "branch", "start_date", "end_date", "target", "delivered", "images"]
NUMERIC_FIELDS = ("target", "delivered", "images", "total_batches", "batches_being_keyed", "promoted")


@dataclass
class MappingResult:
    values: dict = field(default_factory=dict)
    extra: dict = field(default_factory=dict)
    inventory_rows: list = field(default_factory=list)
    errors: list = field(default_factory=list)

    @property
    def is_valid(self):
        return not self.errors


class _SheetCache:
    """Per-import cache so a template with multiple rules pointing at the
    same sheet (e.g. several extra_fields all on "Project Summary") only
    triggers one pandas parse of that sheet, not one per rule."""

    def __init__(self, xls_path):
        self.xls_path = xls_path
        self._frames = {}

    def get(self, sheet_name, header_row=1):
        """header_row is 1-indexed (matches how a person reads row numbers
        in Excel) - a sheet whose real column headers sit on row 2 (e.g. a
        row 1 with target/summary numbers above the header) passes
        header_row=2. Defaults to 1 (pandas header=0), same as before this
        option existed, so every template written before this change keeps
        working unchanged."""
        key = (sheet_name, header_row)
        if key not in self._frames:
            self._frames[key] = pd.read_excel(self.xls_path, sheet_name=sheet_name, header=header_row - 1)
        return self._frames[key]


def _get_cell(wb, sheet, cell):
    ws = wb[sheet]
    return ws[cell].value


def _normalize_header(s):
    """Real Excel headers routinely carry embedded line breaks or double
    spaces from wrapped cells (e.g. "Approx.\\n Records") - config files are
    typically typed by hand without them ("Approx. Records"). A plain
    .strip() only trims the edges, so those two never matched and the
    mapped field silently stayed empty/zero for that column. Collapsing
    every run of whitespace (including embedded newlines) to one space
    before comparing fixes that for every header match in this module."""
    return " ".join(str(s).split()).lower()


def _sum_column(cache, sheet, column, header_row=1):
    df = cache.get(sheet, header_row)
    match = next((c for c in df.columns if _normalize_header(c) == _normalize_header(column)), None)
    if match is None:
        raise KeyError(f"Column '{column}' not found on sheet '{sheet}'")
    # Trackers commonly end a flat sheet with a "Grand Total" row that has a
    # SUM formula's *value* sitting in this same column, but with nothing in
    # the sheet's own leading S.No/serial column - every real data row has
    # that running number, only a trailing total row leaves it blank.
    # Summing that row on top of the real rows silently doubles the result,
    # so it's excluded here using that same signal.
    first_col = df.columns[0]
    if first_col != match:
        df = df[df[first_col].notna()]
    return pd.to_numeric(df[match], errors="coerce").fillna(0).sum()


def _header_lookup(cache, sheet, column, row=0, header_row=1):
    df = cache.get(sheet, header_row)
    matches = [c for c in df.columns if _normalize_header(c) == _normalize_header(column)]
    if not matches:
        raise KeyError(f"Header '{column}' not found on sheet '{sheet}'")
    series = df[matches[0]].dropna()
    if series.empty:
        return None
    return series.iloc[row]


def _row_count(cache, sheet, id_column=None, where_column=None, where_column_blank=None, header_row=1):
    """Counts real data rows on a flat sheet - used for fields like
    `total_batches`/`promoted`/`batches_being_keyed` when the true count of
    batches lives on a raw per-row sheet rather than a single summary cell
    (e.g. a project whose "Inventory" tracking is split across several
    sheets).

    `id_column`, if given, is the header used to identify a genuine row
    (defaults to the sheet's first column, same "leading serial column"
    signal `_sum_column` uses to skip a trailing Grand Total row).
    `where_column`, if given, additionally requires that column to be
    non-blank - e.g. counting only rows whose "Shipment" column is filled
    in, to get a "batches shipped" count instead of "all batches".
    `where_column_blank` is the inverse - counts only rows where that
    column IS blank, e.g. "batches still being keyed" = rows with no
    Shipment date yet. Only one of the two should be set."""
    df = cache.get(sheet, header_row)
    if id_column:
        match = next((c for c in df.columns if _normalize_header(c) == _normalize_header(id_column)), None)
        if match is None:
            raise KeyError(f"Column '{id_column}' not found on sheet '{sheet}'")
    else:
        match = df.columns[0]
    mask = df[match].notna()
    if where_column:
        wmatch = next((c for c in df.columns if _normalize_header(c) == _normalize_header(where_column)), None)
        if wmatch is None:
            raise KeyError(f"Column '{where_column}' not found on sheet '{sheet}'")
        mask &= df[wmatch].notna()
    if where_column_blank:
        bmatch = next((c for c in df.columns if _normalize_header(c) == _normalize_header(where_column_blank)), None)
        if bmatch is None:
            raise KeyError(f"Column '{where_column_blank}' not found on sheet '{sheet}'")
        mask &= df[bmatch].isna()
    return int(mask.sum())


def _coerce_date(value):
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value if isinstance(value, datetime.date) and not isinstance(value, datetime.datetime) else value.date()
    if isinstance(value, str):
        try:
            return pd.to_datetime(value).date()
        except Exception:
            return None
    return None


def extract_field(wb, cache, rule: dict):
    mode = rule.get("mode", "static")
    if mode == "static":
        return rule.get("value")
    if mode == "cell":
        return _get_cell(wb, rule["sheet"], rule["cell"])
    if mode == "sum_column":
        return _sum_column(cache, rule["sheet"], rule["column"], rule.get("header_row", 1))
    if mode == "header":
        return _header_lookup(cache, rule["sheet"], rule["column"], rule.get("row", 0), rule.get("header_row", 1))
    if mode == "row_count":
        return _row_count(cache, rule["sheet"], rule.get("id_column"), rule.get("where_column"), rule.get("where_column_blank"), rule.get("header_row", 1))
    if mode == "sum_multi":
        # Combines the SAME logical value across several sheets - e.g. a
        # project whose batches are split across two Inventory sheets
        # (BV_Inventory_01 + BV_Inventory_02) with no single sheet holding
        # the true total. `sources` is a list of ordinary rules (any mode
        # above, "sum_multi" itself not allowed to avoid infinite nesting);
        # their extracted values are summed. Never hardcodes which/how many
        # sheets - purely driven by whatever `sources` the template lists,
        # so adding a 3rd/4th split sheet for a future project is just
        # another entry in this list, not a code change.
        total = 0
        for sub_rule in rule.get("sources", []):
            total += extract_field(wb, cache, sub_rule) or 0
        return total
    raise ValueError(f"Unknown mapping mode: {mode}")


def _validate_rule_sheet(field_name, rule, sheet_names, errors):
    mode = rule.get("mode", "static")
    if mode == "sum_multi":
        for sub_rule in rule.get("sources", []):
            _validate_rule_sheet(field_name, sub_rule, sheet_names, errors)
        return
    sheet = rule.get("sheet")
    if mode in ("cell", "sum_column", "header", "row_count") and sheet not in sheet_names:
        errors.append(f"[{field_name}] sheet '{sheet}' not found in workbook")


def validate_headers(sheet_names, config: dict) -> list:
    """Pre-flight check: make sure every sheet this template expects
    actually exists in the uploaded file, before we try to import.
    Takes the already-loaded sheet name list rather than the file path, so
    this no longer triggers its own separate workbook parse."""
    errors = []
    for field_name, rule in config.items():
        rules = rule if isinstance(rule, list) else [rule]
        for r in rules:
            if not isinstance(r, dict):
                continue
            if field_name == "inventory_rows":
                sheet = r.get("sheet")
                if sheet and sheet not in sheet_names:
                    errors.append(f"[inventory_rows] sheet '{sheet}' not found in workbook")
            else:
                _validate_rule_sheet(field_name, r, sheet_names, errors)
    return errors


def extract_inventory_rows(cache, rule: dict) -> list:
    """Reads every row of a flat sheet into a list of dicts, using `columns`
    to map standard InventoryItem fields to that sheet's actual header names.
    rule shape: {"sheet": "...", "columns": {"file_name": "File Name", ...}}
    Any column not found is silently skipped (not every project's sheet has
    every field) - this only ever adds detail, never blocks the core import.
    """
    sheet = rule.get("sheet")
    columns = rule.get("columns", {})
    if not sheet or not columns:
        return []

    df = cache.get(sheet, rule.get("header_row", 1))

    # doing this per-row was O(rows x columns) and could take minutes on
    # large inventory sheets (tens of thousands of rows).
    resolved = {}
    for field_name, header in columns.items():
        match = next((c for c in df.columns if _normalize_header(c) == _normalize_header(header)), None)
        if match is not None:
            resolved[field_name] = match
            
    # Auto-detect any other columns present on the sheet so they can be
    # pulled into the `extra` JSON and customized in the Inventory Tracker UI.
    for c in df.columns:
        if c not in resolved.values() and not str(c).startswith("Unnamed:"):
            # Use the literal header name as the key for extra data
            header_str = str(c).strip()
            if header_str:
                resolved[header_str] = c

    if not resolved:
        return []

    subset = df[list(resolved.values())].copy()
    subset.columns = list(resolved.keys())
    subset = subset.astype(object)
    subset[subset.isna()] = None

    records = subset.to_dict("records")

    # Real trackers' raw row range often includes a repeated header row
    # (two-row headers, merged cells) and/or a trailing "Total"/"Grand Total"
    # row with a SUM formula's *value*. Neither is a real inventory item -
    # summing them on top of the individual rows silently double-counts
    # Delivered/Received Records, so they're dropped here before anything
    # downstream (InventoryItem rows, column sums) ever sees them.
    TOTAL_KEYWORDS = {"total", "grand total", "totals", "sub total", "subtotal", "sum"}
    header_labels = {_normalize_header(v) for v in resolved.values()}
    # Whether this template maps a real per-row identifier at all - only
    # templates that do can have this next check applied, so it never
    # misfires on a template that genuinely doesn't track file/folder names.
    has_identifier_column = "file_name" in resolved or "folder_name" in resolved

    def _is_junk_row(row):
        for id_key in ("file_name", "folder_name"):
            val = row.get(id_key)
            if val and str(val).strip().lower() in TOTAL_KEYWORDS:
                return True
        text_vals = {_normalize_header(v) for v in row.values() if isinstance(v, str) and str(v).strip()}
        if text_vals and text_vals.issubset(header_labels):
            return True  # every text cell in this row is literally a column header - a repeated header row
        # A row with NO identifier at all (blank file_name AND blank
        # folder_name) but real numbers elsewhere is a sheet-level "Grand
        # Total" row - a SUM formula's value sitting in the same columns as
        # the real rows, with nothing typed into the ID column. It isn't
        # labeled "Total" so the check above misses it, but every genuine
        # inventory row always has an identifier, so blank-identifier-with-
        # data is itself the signal.
        if has_identifier_column and row.get("file_name") is None and row.get("folder_name") is None:
            return True
        return False

    return [r for r in records if not _is_junk_row(r)]


def safe_load_workbook(xls_path, data_only=True):
    try:
        return openpyxl.load_workbook(xls_path, data_only=data_only, read_only=True)
    except Exception:
        return openpyxl.load_workbook(xls_path, data_only=data_only, read_only=False)


def apply_mapping(xls_path, config: dict) -> MappingResult:
    """Run every field's extraction rule against the uploaded workbook and
    return a standard-schema dict, ready to save into Project.

    config may include an "extra_fields" key: {label: rule, ...}. These are
    extracted the same way but returned under result.extra (not the standard
    schema) - Project Insights, not the main dashboard, is where they show up.
    """
    result = MappingResult()
    cache = _SheetCache(xls_path)

    # Values-only workbook load with read_only=True fallback to read_only=False
    # to prevent openpyxl's streaming dimension parser bug on exported Google Sheets.
    wb = safe_load_workbook(xls_path, data_only=True)
    try:
        sheet_names = wb.sheetnames

        core_config = {k: v for k, v in config.items() if k != "extra_fields"}
        header_errors = validate_headers(sheet_names, core_config)
        if header_errors:
            result.errors.extend(header_errors)
            return result

        for field_name in STANDARD_SCHEMA:
            rule = config.get(field_name)
            if rule is None:
                if field_name in REQUIRED_FIELDS:
                    result.errors.append(f"Mapping is missing a rule for required field '{field_name}'")
                continue
            try:
                raw = extract_field(wb, cache, rule)
            except Exception as exc:
                result.errors.append(f"[{field_name}] {exc}")
                continue

            if field_name in ("start_date", "end_date"):
                raw = _coerce_date(raw)
                if raw is None and field_name in REQUIRED_FIELDS:
                    result.errors.append(f"[{field_name}] could not parse a valid date")
            elif field_name in NUMERIC_FIELDS:
                try:
                    raw = float(raw) if raw not in (None, "") else 0
                except (TypeError, ValueError):
                    raw = 0

            result.values[field_name] = raw

        for req in REQUIRED_FIELDS:
            if result.values.get(req) in (None, ""):
                result.errors.append(f"Required field '{req}' resolved to an empty value")

        extra_config = config.get("extra_fields", {})
        for label, rule in extra_config.items():
            try:
                value = extract_field(wb, cache, rule)
            except Exception as exc:
                value = f"(unavailable: {exc})"
            if isinstance(value, (datetime.date, datetime.datetime)):
                value = str(value)
            result.extra[label] = value

        inventory_rule = config.get("inventory_rows")
        if inventory_rule:
            # Accepts either one rule (single sheet, the common case) or a
            # list of rules - a project whose Inventory tracking is split
            # across several sheets (e.g. BV_Inventory_01 + BV_Inventory_02,
            # each with its own column names) lists one rule per sheet and
            # every sheet's rows are unioned into this project's Inventory
            # Tracker. Adding project #N with a 3-way split is just a 3rd
            # entry in this list - never a code change.
            rules = inventory_rule if isinstance(inventory_rule, list) else [inventory_rule]
            rows = []
            for rule in rules:
                try:
                    rows.extend(extract_inventory_rows(cache, rule))
                except Exception:
                    # inventory rows are enrichment, not required - a failure
                    # on one sheet never blocks the core Project import or
                    # the other sheets' rows.
                    continue
            result.inventory_rows = rows
    finally:
        wb.close()

    return result