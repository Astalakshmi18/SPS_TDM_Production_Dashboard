"""
Auto-Detect Template Mapping
=============================
Given a brand-new project's Excel file, guess a mapping config instead of
requiring someone to hand-write JSON. Three detection strategies, tried in order:

1. **PM02 standard template match** — every real file analyzed so far (BPW,
   Latvia, Newspaper, BV_AT) shares an identical "Project Summary" +
   "Version Control" cell layout. If those sheets exist with the expected
   label cells, we already know exactly where everything is - this is a
   high-confidence match, free and instant, and reuses the same `cell`
   rules as the hand-built templates.

2. **AI-assisted detection (Gemini)** — for anything that doesn't match the
   standard template, if a GEMINI_API_KEY is configured, a preview of every
   sheet's first ~15 rows is sent to Gemini, which understands the semantic
   meaning of headers/labels far better than keyword matching can (e.g. it
   can tell "Volume" means target records from context, not just a literal
   string match). Falls through to keyword scanning if no key is set, the
   call fails, or the response isn't usable JSON - this is always an
   enhancement, never a hard dependency.

3. **Header keyword scan (fallback)** — scans every sheet's header row for
   columns whose names suggest a standard-schema field (e.g. a header
   containing both "target" and "record" -> `target`). Lower confidence;
   the caller should show the result for human review before saving.

Either way, this also tries to guess an `images` source (a "Calculation" or
"Inventory"-like sheet with an image-count column) and an `inventory_rows`
block, so a detected template is immediately useful, not just the 5 core
fields.
"""
import json

import openpyxl
import pandas as pd
from django.conf import settings

from .engine import REQUIRED_FIELDS, STANDARD_SCHEMA

FIELD_KEYWORDS = {
    "project_name": [["project", "name"], ["version", "control"]],
    "start_date": [["start", "date"]],
    "end_date": [["end", "date"]],
    "target": [["target"], ["volume"], ["planned"]],
    "delivered": [["delivered"]],
    "images": [["image"], ["# images"]],
    "language": [["language"]],
    "vendor": [["vendor"], ["indexing", "partner"]],
    "event_type": [["event", "type"]],
    "ocr_status": [["ocr"]],
}

INVENTORY_KEYWORDS = {
    "file_name": [["file", "name"]],
    "folder_name": [["folder", "name"]],
    "event_type": [["event", "type"]],
    "language": [["language"]],
    "image_count": [["image"]],
    "record_count": [["rec"]],
    "shipment_date": [["shipment", "date"], ["received", "date"]],
}


def _build_sheet_previews(xls_path, sheet_names, max_rows=15):
    """First ~15 rows of every sheet, keyed by 1-indexed row -> Excel column
    letter -> value, so an LLM sees the same coordinates a human would."""
    previews = {}
    for sheet_name in sheet_names:
        try:
            df = pd.read_excel(xls_path, sheet_name=sheet_name, header=None, nrows=max_rows).fillna("")
            cols = []
            for i in range(len(df.columns)):
                name, n = "", i
                while n >= 0:
                    name = chr(n % 26 + 65) + name
                    n = n // 26 - 1
                cols.append(name)
            df.columns = cols
            df.index = df.index + 1
            previews[sheet_name] = df.to_dict(orient="index")
        except Exception:
            continue
    return previews


def _ai_detect(xls_path, sheet_names) -> dict | None:
    """Sends a preview of every sheet to Gemini and asks it to map our
    standard schema fields to cell/column locations. Returns None (never
    raises) on any failure - missing key, network error, unparseable
    response - so this is always a pure enhancement over the keyword
    fallback, never a hard dependency for auto-detect to work at all."""
    api_key = getattr(settings, "GEMINI_API_KEY", "")
    if not api_key:
        return None

    try:
        from google import genai

        client = genai.Client(api_key=api_key)
        previews = _build_sheet_previews(xls_path, sheet_names)

        prompt = f"""We have an Excel file. Here is a preview of its sheets (up to the first 15 rows).
The data is represented as a dictionary where keys are Excel row numbers (1, 2, 3...) and values are
dictionaries of column letters (A, B, C...) to cell values.

{json.dumps(previews, indent=2, default=str)}

Map the fields to our required standard schema:
{json.dumps(STANDARD_SCHEMA, indent=2)}

You can use these mapping modes:
1. "cell": a specific fixed cell (e.g., Project Name is always in cell B1).
   Format: {{"mode": "cell", "sheet": "SheetName", "cell": "B1"}}
2. "sum_column": sum all numeric values in a column with a specific header on a tabular sheet (e.g., Total Images).
   Format: {{"mode": "sum_column", "sheet": "SheetName", "column": "HeaderName"}}
3. "header": find a column header on a tabular sheet and get its first value.
   Format: {{"mode": "header", "sheet": "SheetName", "column": "HeaderName"}}
4. "static": a hardcoded value (e.g. branch is always "TDM" unless the file suggests otherwise).
   Format: {{"mode": "static", "value": "TDM"}}

Conventions:
- "project_name", "start_date", "end_date", "target", "delivered" are usually fixed cells ("cell" mode)
  on a summary or version-control sheet.
- "images" is usually a "sum_column" on a calculation or inventory sheet.
- Always map "branch" to {{"mode": "static", "value": "TDM"}} unless the file clearly states otherwise.

Return ONLY a raw JSON object (no markdown, no code fences) where keys are standard schema fields and
values are the mapping rule dicts above. Only include a field if there's a genuinely probable match -
omit anything you're not confident about rather than guessing.
"""
        response = client.models.generate_content(model="gemini-1.5-flash", contents=prompt)
        text = response.text.strip()
        if text.startswith("```json"):
            text = text[7:]
        elif text.startswith("```"):
            text = text[3:]
        if text.endswith("```"):
            text = text[:-3]

        ai_config = json.loads(text.strip())
        return ai_config if isinstance(ai_config, dict) and ai_config else None
    except Exception:
        return None


def _matches(header: str, keyword_groups: list) -> bool:
    header_lower = str(header).strip().lower()
    return any(all(kw in header_lower for kw in group) for group in keyword_groups)


def _is_pm02_standard(wb) -> bool:
    if "Project Summary" not in wb.sheetnames or "Version Control" not in wb.sheetnames:
        return False
    ws = wb["Project Summary"]
    e1 = str(ws["E1"].value or "").lower()
    g1 = str(ws["G1"].value or "").lower()
    return "start" in e1 and "date" in e1 and "end" in g1


def _detect_images_source(wb):
    """Look for a sheet whose header row has a column mentioning 'image',
    preferring sheets named like Calculation/Inventory."""
    candidates = sorted(wb.sheetnames, key=lambda s: 0 if s.lower() in ("calculation", "inventory") else 1)
    for sheet_name in candidates:
        ws = wb[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        for cell in header_row:
            if cell and "image" in str(cell).lower():
                return {"sheet": sheet_name, "column": str(cell)}
    return None


def _detect_inventory_rows(wb):
    candidates = sorted(wb.sheetnames, key=lambda s: 0 if s.lower() in ("calculation", "inventory") else 1)
    for sheet_name in candidates:
        ws = wb[sheet_name]
        header_row = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []) if c]
        if not header_row:
            continue
        columns = {}
        for field_name, groups in INVENTORY_KEYWORDS.items():
            match = next((h for h in header_row if _matches(h, groups)), None)
            if match:
                columns[field_name] = match
        if len(columns) >= 2:  # only worth it if we found meaningful structure
            return {"sheet": sheet_name, "columns": columns}
    return None


def detect_mapping(xls_path: str) -> dict:
    """Returns {"config": {...}, "notes": [...], "confidence": "high"|"medium"|"low"}"""
    try:
        wb = openpyxl.load_workbook(xls_path, data_only=True, read_only=True)
    except Exception:
        wb = openpyxl.load_workbook(xls_path, data_only=True, read_only=False)

    try:
        return _detect_mapping_inner(xls_path, wb)
    finally:
        wb.close()


def _detect_mapping_inner(xls_path: str, wb) -> dict:
    notes = []

    if _is_pm02_standard(wb):
        config = {
            "project_name": {"mode": "cell", "sheet": "Version Control", "cell": "B1"},
            "branch": {"mode": "static", "value": "TDM"},
            "start_date": {"mode": "cell", "sheet": "Project Summary", "cell": "D1"},
            "end_date": {"mode": "cell", "sheet": "Project Summary", "cell": "F1"},
            "target": {"mode": "cell", "sheet": "Project Summary", "cell": "B2"},
            "delivered": {"mode": "cell", "sheet": "Project Summary", "cell": "B4"},
        }
        notes.append("Matched the standard PM02 'Project Summary' + 'Version Control' layout (high confidence).")
        confidence = "high"

        images = _detect_images_source(wb)
        if images:
            config["images"] = {"mode": "sum_column", "sheet": images["sheet"], "column": images["column"]}
            notes.append(f"Found an images column '{images['column']}' on sheet '{images['sheet']}'.")
        else:
            notes.append("Could not find an images column automatically - add an 'images' rule manually before saving.")

        inv = _detect_inventory_rows(wb)
        if inv:
            config["inventory_rows"] = inv
            notes.append(f"Detected inventory row detail on sheet '{inv['sheet']}' ({len(inv['columns'])} columns matched).")

        return {"config": config, "notes": notes, "confidence": confidence}

    # Tier 2: AI-assisted detection (only if GEMINI_API_KEY is configured).
    ai_config = _ai_detect(xls_path, wb.sheetnames)
    if ai_config:
        notes.append("Matched via AI-assisted detection (Gemini) - review carefully, this is semantic guessing, not a proven layout.")
        confidence = "medium"

        if "branch" not in ai_config:
            ai_config["branch"] = {"mode": "static", "value": "TDM"}

        inv = _detect_inventory_rows(wb)
        if inv and "inventory_rows" not in ai_config:
            ai_config["inventory_rows"] = inv
            notes.append(f"Detected inventory row detail on sheet '{inv['sheet']}' ({len(inv['columns'])} columns matched).")

        missing = set(REQUIRED_FIELDS) - set(ai_config.keys())
        if missing:
            notes.append(f"AI did not confidently map: {', '.join(sorted(missing))}. Fill these in manually before saving.")
            confidence = "low"

        return {"config": ai_config, "notes": notes, "confidence": confidence}

    # Tier 3: keyword-scan every sheet's header row.
    if getattr(settings, "GEMINI_API_KEY", ""):
        notes.append("AI-assisted detection didn't return a usable result - fell back to scanning column headers. Review carefully.")
    else:
        notes.append("Did not match the standard PM02 layout - fell back to scanning column headers. Review carefully before saving. "
                      "(Set GEMINI_API_KEY to enable smarter AI-assisted detection for non-standard formats.)")
    config = {"branch": {"mode": "static", "value": "TDM"}}
    found_fields = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []) if c]
        if not header_row:
            continue
        for field_name, groups in FIELD_KEYWORDS.items():
            if field_name in found_fields or field_name == "branch":
                continue
            match = next((h for h in header_row if _matches(h, groups)), None)
            if match:
                config[field_name] = {"mode": "header", "sheet": sheet_name, "column": str(match)}
                found_fields.add(field_name)
                notes.append(f"Guessed '{field_name}' -> sheet '{sheet_name}', column '{match}'.")

    required = {"project_name", "start_date", "end_date", "target", "delivered", "images"}
    missing = required - found_fields
    if missing:
        notes.append(f"Could not confidently detect: {', '.join(sorted(missing))}. Fill these in manually before saving.")
        confidence = "low"
    else:
        confidence = "medium"

    inv = _detect_inventory_rows(wb)
    if inv:
        config["inventory_rows"] = inv
        notes.append(f"Detected inventory row detail on sheet '{inv['sheet']}' ({len(inv['columns'])} columns matched).")

    return {"config": config, "notes": notes, "confidence": confidence}
