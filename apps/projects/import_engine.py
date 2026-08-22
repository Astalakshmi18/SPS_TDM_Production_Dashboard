"""
Excel Import Engine
====================
Workflow (matches the brief exactly):
  1. User uploads Excel.
  2. Select project (ProjectTemplate).
  3. Load mapping (ProjectTemplate.config).
  4. Validate headers (mapping.engine.validate_headers).
  5. Convert into common schema (mapping.engine.apply_mapping).
  6. Save into database (Project, upserted by project_key + branch).
  7. Dashboard auto refreshes (it always reads live from Project - no cache).
"""
from django.db import transaction

from apps.branches.models import Branch
from apps.mapping.engine import apply_mapping
from .models import ImportBatch, Project


def _cell_num(ws, addr):
    """Reads a numeric cell defensively - real sheets have stray text
    ("Exatech" typed into a Received Records cell) and Excel errors
    (#VALUE!) sitting in cells that are supposed to be numbers. Either
    just means "not available", not a crash."""
    v = ws[addr].value
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        try:
            return float(s.replace(",", "").replace("%", ""))
        except ValueError:
            return None
    if hasattr(v, "year"):  # a date landed in a numeric cell - not usable here
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _cell_date(ws, addr):
    v = ws[addr].value
    return v.date() if hasattr(v, "date") else (v if hasattr(v, "year") else None)


def read_project_summary_snapshot(xls_path):
    """Reads the Project Summary sheet's OWN pre-computed cells - Delivered/
    Received Records, %, Days Gone, Remaining Days etc. - directly, instead
    of Python re-deriving them. Every real project file checked (Latvia,
    Newspaper, BPW, BV_AT) uses the identical cell layout:
      B1/D1/F1/H1  = status-as-of date / start / end / total days
      B2/C2        = volume / unit label ("records" or "Pages")
      A4/B4/C4     = Delivered label / value / Days Gone
      B5           = Delivered %
      B6           = Days Gone %
      A8/B8/C8     = Remaining label / value / Remaining Days
      B9           = Remaining %
      B10          = Remaining Days %
      B13/A14/B14  = Branch Status as-of date / Received label / value
      B15          = Received %
    Sanity-checked against A1's label before trusting any of it - if a
    project's sheet doesn't follow this layout, this quietly returns {}
    and the panel falls back to the Selected_Column / imported-field path."""
    import openpyxl

    try:
        try:
            wb = openpyxl.load_workbook(xls_path, data_only=True, read_only=True)
        except Exception:
            wb = openpyxl.load_workbook(xls_path, data_only=True, read_only=False)
    except Exception:
        return {}
    if "Project Summary" not in wb.sheetnames:
        return {}
    ws = wb["Project Summary"]

    a1 = str(ws["A1"].value or "").strip().lower()
    if "project status" not in a1:
        return {}  # not the layout we expect - don't guess

    delivered_pct = _cell_num(ws, "B5")
    days_gone_pct = _cell_num(ws, "B6")
    remaining_pct = _cell_num(ws, "B9")
    remaining_days_pct = _cell_num(ws, "B10")
    received_pct = _cell_num(ws, "B15")

    return {
        "status_as_of": str(_cell_date(ws, "B1") or ""),
        "start_date": str(_cell_date(ws, "D1") or ""),
        "end_date": str(_cell_date(ws, "F1") or ""),
        "total_days": _cell_num(ws, "H1"),
        "volume": _cell_num(ws, "B2"),
        "volume_unit": str(ws["C2"].value or "").strip(),
        "delivered_label": str(ws["A4"].value or "Delivered Records").strip(),
        "delivered": _cell_num(ws, "B4"),
        "days_gone": _cell_num(ws, "C4"),
        "delivered_pct": round(delivered_pct * 100, 2) if delivered_pct is not None else None,
        "days_gone_pct": round(days_gone_pct * 100, 2) if days_gone_pct is not None else None,
        "remaining_label": str(ws["A8"].value or "Remaining Records").strip(),
        "remaining": _cell_num(ws, "B8"),
        "remaining_days": _cell_num(ws, "C8"),
        "remaining_pct": round(remaining_pct * 100, 2) if remaining_pct is not None else None,
        "remaining_days_pct": round(remaining_days_pct * 100, 2) if remaining_days_pct is not None else None,
        "branch_status_as_of": str(_cell_date(ws, "B13") or ""),
        "received_label": str(ws["A14"].value or "Received Records").strip(),
        "received": _cell_num(ws, "B14"),
        "received_pct": round(received_pct * 100, 2) if received_pct is not None else None,
    }


def _safe_int(value):
    """Coerces an Excel cell value to int, tolerating the placeholder junk
    real trackers are full of: '-' or '—' for zero, blank strings, 'N/A',
    thousands separators ('1,234'), stray whitespace, floats-as-strings,
    and outright None. Anything that isn't actually a number becomes 0
    rather than crashing the whole import over one bad cell."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        try:
            import math
            if isinstance(value, float) and math.isnan(value):
                return 0
        except Exception:
            pass
        return int(value)
    text = str(value).strip().replace(",", "")
    if text in ("", "-", "—", "–", "N/A", "n/a", "NA", "None"):
        return 0
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return 0


def run_import(template, uploaded_file, django_file_field_path, user,
                source_type=ImportBatch.SOURCE_FILE, source_url=""):
    """template: ProjectTemplate instance
    uploaded_file: the saved path on disk (import needs a real path for pandas/openpyxl) -
        for a Google Sheet import this is the downloaded .xlsx copy, not the live sheet.
    django_file_field_path: original filename (or sheet title), for the audit log
    user: request.user
    source_type / source_url: recorded on ImportBatch for traceability
    """
    result = apply_mapping(uploaded_file, template.config)

    if not result.is_valid:
        ImportBatch.objects.create(
            project_template_key=template.project_key,
            file_name=django_file_field_path,
            source_type=source_type,
            source_url=source_url,
            uploaded_by=user,
            status=ImportBatch.STATUS_FAILED,
            errors=result.errors,
        )
        return None, result.errors

    values = result.values
    branch = Branch.objects.filter(code=values.get("branch")).first() or template.branch

    defaults = {
        "project_name": values.get("project_name") or template.display_name,
        "start_date": values["start_date"],
        "end_date": values["end_date"],
        "target_records": _safe_int(values.get("target")),
        "delivered_records": _safe_int(values.get("delivered")),
        "total_images": _safe_int(values.get("images")),
        "total_batches": _safe_int(values.get("total_batches")),
        "batches_being_keyed": _safe_int(values.get("batches_being_keyed")),
        "promoted": _safe_int(values.get("promoted")),
        "language": values.get("language") or "",
        "vendor": values.get("vendor") or "",
        "event_type": values.get("event_type") or "",
        "ocr_status": values.get("ocr_status") or "",
        "extra_data": result.extra,
        "summary_snapshot": read_project_summary_snapshot(uploaded_file),
    }
    if source_type == ImportBatch.SOURCE_GOOGLE_SHEET and source_url:
        defaults["google_sheet_url"] = source_url

    project, _ = Project.objects.update_or_create(
        project_key=template.project_key,
        branch=branch,
        defaults=defaults,
    )

    # Everything from here is one commit, not several - the Project upsert,
    # the full inventory replace, and the derived batch metrics used to each
    # autocommit separately, which is real added time on a big sync.
    with transaction.atomic():
        _replace_inventory_items(project, result.inventory_rows)

        # Only derive batch metrics from inventory rows when the template
        # didn't already explicitly map total_batches/promoted to real
        # cells/columns - an explicit mapping is more authoritative than our
        # own row-count guess and must never be silently overwritten.
        if values.get("total_batches") in (None, "") and values.get("promoted") in (None, ""):
            _derive_batch_metrics(project)

    ImportBatch.objects.create(
        project_template_key=template.project_key,
        file_name=django_file_field_path,
        source_type=source_type,
        source_url=source_url,
        uploaded_by=user,
        status=ImportBatch.STATUS_SUCCESS,
        project=project,
    )

    return project, []


def _clean_text(value):
    """Converts a raw cell value to display text, stripping the trailing
    '.0' that whole-number-valued file/folder names pick up when Excel
    stores them as floats (e.g. 105913085.0 -> "105913085", not left as-is
    which would show a meaningless decimal on every ID in the Inventory
    Tracker). Also treats NaN as blank - pandas represents an empty Excel
    cell as float('nan'), and str(nan) is the truthy string "nan", which
    would otherwise defeat blank-row detection entirely."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value != value:  # NaN != NaN is the classic, dependency-free NaN check
            return ""
        if value.is_integer():
            return str(int(value))
    return str(value)


def _replace_inventory_items(project, rows):
    """Every (re)import fully replaces this project's inventory rows - the
    source file/sheet is always the single source of truth, so partial
    merges would just accumulate stale rows over time.

    Wrapped in one DB transaction with a chunked bulk_create: on a project
    with tens of thousands of inventory rows, letting Django autocommit each
    statement (the default) was the single biggest cost in "Sync Now" -
    batching writes into one transaction cuts that dramatically."""
    from django.db import transaction

    from apps.inventory.models import InventoryItem

    if not rows:
        return

    items = []
    for row in rows:
        shipment_date = row.get("shipment_date")
        if hasattr(shipment_date, "date"):
            shipment_date = shipment_date.date()
        elif shipment_date is not None and not hasattr(shipment_date, "year"):
            shipment_date = None

        file_name = _clean_text(row.get("file_name"))
        folder_name = _clean_text(row.get("folder_name"))
        image_count = _safe_int(row.get("image_count"))
        record_count = _safe_int(row.get("record_count"))

        # Skip fully-blank rows (e.g. an unfinished project's Calculation
        # sheet with placeholder empty rows) - these add nothing but noise
        # to the Inventory Tracker.
        if not file_name and not folder_name and not image_count and not record_count and not shipment_date:
            continue

        known = {"file_name", "folder_name", "event_type", "language",
                 "image_count", "record_count", "shipment_date", "remarks"}
        extra = {
            k: (v.strftime("%Y-%m-%d") if hasattr(v, "year") else v)
            for k, v in row.items() if k not in known
        }

        items.append(InventoryItem(
            project=project,
            file_name=file_name,
            folder_name=folder_name,
            event_type=_clean_text(row.get("event_type")),
            language=_clean_text(row.get("language")),
            image_count=image_count,
            record_count=record_count,
            shipment_date=shipment_date,
            remarks=str(row.get("remarks") or ""),
            extra=extra,
        ))

    with transaction.atomic():
        InventoryItem.objects.filter(project=project).delete()
        InventoryItem.objects.bulk_create(items, batch_size=1000)


def _derive_batch_metrics(project):
    """Populates total_batches / batches_being_keyed / promoted from the
    project's actual InventoryItem rows, so these are genuinely distinct
    numbers from delivered_records - not a copy of it. A "batch" here means
    one inventory row (one file/folder); "promoted" means that batch has a
    shipment_date recorded (it's shipped out, same concept as PHX's
    "Promoted"), vs "delivered_records" which counts individual records
    within those files. Different units, different numbers, on purpose:
    a project can easily have e.g. 8 batches (files) worth 69,548 delivered
    records - both are correct, they're just not the same measurement.

    One aggregate query (not two separate .count() table scans) for speed
    on large inventories."""
    from django.db.models import Count, Q

    from apps.inventory.models import InventoryItem

    agg = InventoryItem.objects.filter(project=project).aggregate(
        total=Count("id"),
        promoted=Count("id", filter=Q(shipment_date__isnull=False)),
    )
    total_batches = agg["total"] or 0
    if not total_batches:
        return  # no inventory_rows mapped for this template - leave batch fields at 0, not misleadingly equal to anything

    promoted = agg["promoted"] or 0
    Project.objects.filter(pk=project.pk).update(
        total_batches=total_batches,
        promoted=promoted,
        batches_being_keyed=total_batches - promoted,
    )
    project.total_batches = total_batches
    project.promoted = promoted
    project.batches_being_keyed = total_batches - promoted


def resync_project(project, user=None):
    """Re-pull a project's data from wherever it originally came from - used
    by the "Sync Now" button and by the Google Apps Script webhook. Only
    works for projects that were imported from a Google Sheet (google_sheet_url
    is set); returns (project, errors) same shape as run_import."""
    from apps.mapping.models import ProjectTemplate
    from .gsheet import download_as_xlsx

    if not project.google_sheet_url:
        return None, ["This project has no linked Google Sheet to sync from."]

    template = ProjectTemplate.objects.filter(project_key=project.project_key).first()
    if not template:
        return None, [f"No mapping template found for project_key '{project.project_key}'."]

    full_path = download_as_xlsx(project.google_sheet_url)
    return run_import(
        template, full_path, project.google_sheet_url, user,
        source_type=ImportBatch.SOURCE_GOOGLE_SHEET, source_url=project.google_sheet_url,
    )
