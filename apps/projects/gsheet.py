"""
Google Sheet import support.

A project's tracker can live as a Google Sheet instead of an uploaded .xlsx -
same mapping engine, same standard schema, only the source of the bytes
differs. This module resolves any of the common Google Sheets share-link
shapes to the sheet's ID, downloads it as .xlsx, and saves it to disk so the
existing mapping engine (which needs a real file path for pandas/openpyxl)
can read it exactly like an upload.

Auth model: OAuth2 as a specific Google account - NOT "Anyone with the link
can view". The sheet only needs to be shared with one Google account (role:
Viewer is enough), exactly like sharing it with a person. The server signs
in as that account using a long-lived refresh token and calls the Drive API
export endpoint, so it can read anything that account has been given access
to without the sheet ever being publicly link-accessible.

One-time setup to connect an account (see SETUP_GSHEET_OAUTH.md in the repo
root for the full click-by-click steps):
    1. Create an OAuth Client ID (type: Desktop app) in Google Cloud Console.
    2. Using the OAuth Playground (or equivalent), sign in as the Google
       account that was given Viewer access to the sheet(s), and generate a
       refresh token for the scope https://www.googleapis.com/auth/drive.readonly
    3. Set these env vars on the server (Render dashboard -> Environment):
         GOOGLE_OAUTH_CLIENT_ID
         GOOGLE_OAUTH_CLIENT_SECRET
         GOOGLE_OAUTH_REFRESH_TOKEN

Large-sheet fallback: Drive's "export as .xlsx" conversion has its own hard
size ceiling (Google returns 403 exportSizeLimitExceeded, "This file is too
large to be exported" - seen in practice on a multi-thousand-row Inventory
tracker) that has nothing to do with sharing/permissions. When that specific
error is hit, this module rebuilds an equivalent local .xlsx itself using
the Sheets API (per-sheet cell values + number-format-type, to tell real
dates apart from plain numbers) instead of asking Drive to convert the
whole file at once - the same "drive.readonly" scope already covers the
Sheets API too, no re-authorization needed.
"""
import datetime
import re
import uuid

import requests
from django.conf import settings

SHEET_ID_PATTERNS = [
    r"/spreadsheets/d/([a-zA-Z0-9-_]+)",  # standard share link
    r"^([a-zA-Z0-9-_]{20,})$",             # bare sheet ID pasted directly
]

TOKEN_URL = "https://oauth2.googleapis.com/token"
DRIVE_EXPORT_URL = "https://www.googleapis.com/drive/v3/files/{file_id}/export"
SHEETS_API_URL = "https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
GOOGLE_SERIAL_EPOCH = datetime.date(1899, 12, 30)  # matches Excel's own date system


class GoogleSheetError(Exception):
    pass


def extract_sheet_id(url_or_id: str) -> str:
    url_or_id = url_or_id.strip()
    for pattern in SHEET_ID_PATTERNS:
        m = re.search(pattern, url_or_id)
        if m:
            return m.group(1)
    raise GoogleSheetError("Could not find a Google Sheet ID in that link.")


def _get_access_token() -> str:
    """Exchanges the long-lived refresh token for a short-lived access
    token. This authenticates as whichever Google account the refresh
    token belongs to - i.e. the account that was given Viewer access to
    the sheet(s) - so no sheet needs to be publicly link-shared."""
    client_id = getattr(settings, "GOOGLE_OAUTH_CLIENT_ID", "")
    client_secret = getattr(settings, "GOOGLE_OAUTH_CLIENT_SECRET", "")
    refresh_token = getattr(settings, "GOOGLE_OAUTH_REFRESH_TOKEN", "")

    if not (client_id and client_secret and refresh_token):
        raise GoogleSheetError(
            "No Google account is connected yet. Set GOOGLE_OAUTH_CLIENT_ID, "
            "GOOGLE_OAUTH_CLIENT_SECRET and GOOGLE_OAUTH_REFRESH_TOKEN as env "
            "vars - see SETUP_GSHEET_OAUTH.md for the one-time steps."
        )

    try:
        response = requests.post(TOKEN_URL, data={
            "client_id": client_id,
            "client_secret": client_secret,
            "refresh_token": refresh_token,
            "grant_type": "refresh_token",
        }, timeout=15)
    except requests.RequestException as exc:
        raise GoogleSheetError(f"Could not reach Google to refresh the access token: {exc}") from exc

    if response.status_code != 200:
        raise GoogleSheetError(
            f"Google rejected the stored credentials (HTTP {response.status_code}). "
            "The refresh token may have been revoked or the connected account's "
            "password/security settings changed - redo the one-time setup in "
            "SETUP_GSHEET_OAUTH.md."
        )

    token = response.json().get("access_token")
    if not token:
        raise GoogleSheetError("Google did not return an access token.")
    return token


def _is_export_size_limit_error(response) -> bool:
    try:
        reason = response.json().get("error", {}).get("errors", [{}])[0].get("reason", "")
    except Exception:
        reason = ""
    return reason == "exportSizeLimitExceeded" or "too large to be exported" in response.text.lower()


def _sheet_titles(sheet_id, access_token):
    resp = requests.get(
        SHEETS_API_URL.format(sheet_id=sheet_id),
        params={"fields": "sheets.properties.title"},
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=30,
    )
    if resp.status_code != 200:
        raise GoogleSheetError(f"Could not list this sheet's tabs (HTTP {resp.status_code}).")
    return [s["properties"]["title"] for s in resp.json().get("sheets", [])]


def _date_mask_for_sheet(sheet_id, title, access_token):
    """Per-cell True/False grid (same shape as the values grid) saying
    whether that cell is formatted as a date/datetime - values.get alone
    can't tell a date apart from a plain number once both are pulled as raw
    (UNFORMATTED_VALUE) serial numbers, so this is fetched separately.
    Format-only (no cell values in this call), which keeps it much lighter
    than asking for full styled grid data on a huge sheet."""
    resp = requests.get(
        SHEETS_API_URL.format(sheet_id=sheet_id),
        params={
            "ranges": title,
            "fields": "sheets.data.rowData.values.effectiveFormat.numberFormat.type",
        },
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=60,
    )
    if resp.status_code != 200:
        return []  # non-fatal - cells just won't be recognized as dates
    sheets_data = resp.json().get("sheets", [])
    if not sheets_data:
        return []
    row_data = sheets_data[0].get("data", [{}])[0].get("rowData", [])
    mask = []
    for row in row_data:
        row_mask = []
        for cell in row.get("values", []):
            fmt_type = cell.get("effectiveFormat", {}).get("numberFormat", {}).get("type")
            row_mask.append(fmt_type in ("DATE", "DATE_TIME"))
        mask.append(row_mask)
    return mask


def _sheet_values(sheet_id, title, access_token):
    from urllib.parse import quote

    resp = requests.get(
        f"{SHEETS_API_URL.format(sheet_id=sheet_id)}/values/{quote(title, safe='')}",
        params={"valueRenderOption": "UNFORMATTED_VALUE", "dateTimeRenderOption": "SERIAL_NUMBER"},
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=90,
    )
    if resp.status_code != 200:
        return []  # skip this tab rather than fail the whole workbook
    return resp.json().get("values", [])


def _rebuild_via_sheets_api(sheet_id, access_token):
    """Reconstructs an equivalent local .xlsx by reading raw cell data
    through the Sheets API (one sheet/tab at a time) instead of asking
    Drive to convert the whole workbook to .xlsx in one shot - the
    conversion step is what actually has the size ceiling, not reading the
    data itself.

    Known limitation: only DATE/DATE_TIME formatted cells are converted
    back from Google's raw serial-number representation - a cell formatted
    as PERCENT comes back from UNFORMATTED_VALUE as its raw fraction (0.15,
    not 15), unlike the normal Drive-export path where Excel's own percent
    formatting is preserved. Only affects sheets big enough to hit this
    fallback; fine for every field currently mapped from such a sheet
    (Inventory rows), but worth knowing if a future template maps a
    percent-formatted cell from an oversized sheet."""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_titles = set()
    for title in _sheet_titles(sheet_id, access_token):
        # Excel sheet names can't contain \ / ? * [ ] : (Google Sheets tab
        # names have no such restriction, so a real tab title like "BV/AT"
        # would otherwise crash openpyxl here) and are capped at 31 chars.
        safe_title = re.sub(r'[\\/?*\[\]:]', "-", title)[:31].strip() or "Sheet"
        if safe_title in used_titles:
            suffix = 2
            while f"{safe_title[:28]}_{suffix}" in used_titles:
                suffix += 1
            safe_title = f"{safe_title[:28]}_{suffix}"
        used_titles.add(safe_title)
        ws = wb.create_sheet(title=safe_title)

        date_mask = _date_mask_for_sheet(sheet_id, title, access_token)
        values = _sheet_values(sheet_id, title, access_token)

        for r_idx, row in enumerate(values):
            for c_idx, value in enumerate(row):
                is_date = (
                    r_idx < len(date_mask) and c_idx < len(date_mask[r_idx])
                    and date_mask[r_idx][c_idx]
                    and isinstance(value, (int, float))
                )
                cell_value = (GOOGLE_SERIAL_EPOCH + datetime.timedelta(days=value)) if is_date else value
                ws.cell(row=r_idx + 1, column=c_idx + 1, value=cell_value)

    if not wb.sheetnames:
        raise GoogleSheetError("Could not read any tabs from this sheet via the Sheets API either.")
    return wb


def download_as_xlsx(url_or_id: str) -> str:
    """Downloads the sheet (as the connected Google account) and returns
    the local file path it was saved to."""
    sheet_id = extract_sheet_id(url_or_id)
    access_token = _get_access_token()

    try:
        response = requests.get(
            DRIVE_EXPORT_URL.format(file_id=sheet_id),
            params={"mimeType": XLSX_MIME},
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=30,
        )
    except requests.RequestException as exc:
        raise GoogleSheetError(f"Could not reach Google Sheets: {exc}") from exc

    uploads_dir = settings.MEDIA_ROOT / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    file_name = f"gsheet_{sheet_id[:8]}_{uuid.uuid4().hex[:6]}.xlsx"
    file_path = uploads_dir / file_name

    if response.status_code == 403 and _is_export_size_limit_error(response):
        # Drive's own "convert to .xlsx" step has a hard size ceiling this
        # sheet exceeds - rebuild it ourselves via the Sheets API instead
        # (same account, same permission, no re-authorization needed).
        wb = _rebuild_via_sheets_api(sheet_id, access_token)
        wb.save(file_path)
        return str(file_path)

    if response.status_code == 404:
        raise GoogleSheetError(
            "Sheet not found, or the connected Google account doesn't have "
            "access to it. Share the sheet with that account (Viewer is "
            "enough) and try again."
        )
    if response.status_code == 403:
        raise GoogleSheetError(
            "The connected Google account was denied access to this sheet. "
            "Make sure it has at least Viewer permission on it."
        )
    if response.status_code != 200:
        raise GoogleSheetError(f"Google Sheets returned HTTP {response.status_code}.")

    file_path.write_bytes(response.content)
    return str(file_path)
